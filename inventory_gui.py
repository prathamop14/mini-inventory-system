import tkinter as tk
from tkinter import messagebox
import mysql.connector
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import pandas as pd
from datetime import datetime


# Database connection
def create_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="inventory_db"
    )

# Add item function
def add_item():
    name = name_entry.get()
    price = price_entry.get()
    qty = qty_entry.get()

    if not name or not price or not qty:
        messagebox.showerror("Error", "All fields required")
        return

    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO items (name, price, quantity) VALUES (%s,%s,%s)",
        (name, price, qty)
    )
    conn.commit()
    conn.close()

    messagebox.showinfo("Success", "Item Added")
    clear_fields()
    view_items()

# View items
def view_items():
    listbox.delete(0, tk.END)
    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM items")
    rows = cursor.fetchall()
    for row in rows:
        listbox.insert(tk.END, row)
    conn.close()

def export_to_excel():
    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM items")
    rows = cursor.fetchall()
    conn.close()

    df = pd.DataFrame(rows, columns=["ID", "Name", "Price", "Quantity"])
    df["Total Value"] = df["Price"] * df["Quantity"]

    file_name = "inventory_dashboard.xlsx"
    df.to_excel(file_name, index=False, startrow=6)

    wb = load_workbook(file_name)
    ws = wb.active

    # 🔵 Dashboard Title
    ws["A1"] = "INVENTORY DASHBOARD"
    ws["A1"].font = Font(size=18, bold=True)
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center")

    # 🕒 Date Time
    ws["A2"] = "Generated On: " + datetime.now().strftime("%d-%m-%Y %H:%M:%S")

    # 📊 Summary Section
    total_items = len(df)
    total_quantity = df["Quantity"].sum()
    total_value = df["Total Value"].sum()

    ws["A4"] = "Total Products:"
    ws["B4"] = total_items

    ws["C4"] = "Total Quantity:"
    ws["D4"] = total_quantity

    ws["E4"] = "Total Inventory Value:"
    ws["F4"] = total_value

    ws["F4"].number_format = '₹#,##0.00'

    # 🎨 Header Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=7, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 📏 Auto Column Width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 3

    # 🔲 Border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # 📊 Bar Chart (Quantity Chart)
    chart = BarChart()
    chart.title = "Stock Quantity Chart"
    chart.y_axis.title = "Quantity"
    chart.x_axis.title = "Products"

    data = Reference(ws, min_col=4, min_row=7, max_row=7 + total_items)
    cats = Reference(ws, min_col=2, min_row=8, max_row=7 + total_items)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "H6")

    wb.save(file_name)

    messagebox.showinfo("Success", "Dashboard created successfully!")


# Clear
def clear_fields():
    name_entry.delete(0, tk.END)
    price_entry.delete(0, tk.END)
    qty_entry.delete(0, tk.END)

# Window
root = tk.Tk()
root.title("Inventory Management")
root.geometry("600x500")

tk.Label(root, text="Name").pack()
name_entry = tk.Entry(root)
name_entry.pack()

tk.Label(root, text="Price").pack()
price_entry = tk.Entry(root)
price_entry.pack()

tk.Label(root, text="Quantity").pack()
qty_entry = tk.Entry(root)
qty_entry.pack()

tk.Button(root, text="Add Item", command=add_item).pack(pady=5)
tk.Button(root, text="View Items", command=view_items).pack(pady=5)
tk.Button(root, text="Export to Excel", command=export_to_excel).pack(pady=5)

listbox = tk.Listbox(root, width=70)
listbox.pack(pady=10)

view_items()

root.mainloop()
